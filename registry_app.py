import os
import json
import threading
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, unquote
from openpyxl import load_workbook, Workbook
from datetime import datetime

# ─── Конфиг ──────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")

DEFAULT_CONFIG = {
    
}

def load_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return DEFAULT_CONFIG

def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

# ─── Логика чтения ────────────────────────────────────────────────────────────

def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    try:
        return str(value).split(" ")[0]
    except:
        return ""

def read_cell(sheet, cell_addr, col_type):
    try:
        val = sheet[cell_addr].value
    except:
        return None, "Неверный адрес ячейки"
    if val is None:
        return None, None
    if col_type == "date":
        return format_date(val), None
    if col_type == "number":
        try:
            return float(val), None
        except:
            return 0.0, None
    return str(val), None

def build_registry(folder, config):
    columns = config.get("columns", [])
    rows = []
    errors = []

    for filename in sorted(os.listdir(folder)):
        if not filename.endswith((".xlsx", ".xlsm", ".xls")):
            continue
        if filename == "Реестр.xlsx":
            continue

        filepath = os.path.join(folder, filename)
        try:
            wb = load_workbook(filepath, data_only=True)
            sheet = wb.active

            row = {"_file": filename}
            row_errors = []

            for col in columns:
                val, err = read_cell(sheet, col["cell"].upper(), col["type"])
                if err:
                    row_errors.append(f'{col["name"]} ({col["cell"]}): {err}')
                    val = None
                if col.get("required") and (val is None or val == ""):
                    row_errors.append(f'{col["name"]} ({col["cell"]}): пустое обязательное поле')
                row[col["name"]] = val

            rows.append(row)
            if row_errors:
                errors.append({"file": filename, "error": "; ".join(row_errors)})
            else:
                pass  # ok

        except Exception as e:
            errors.append({"file": filename, "error": str(e)})

    return {"rows": rows, "errors": errors, "columns": [c["name"] for c in columns]}

def save_registry_xlsx(folder, rows, config):
    columns = config.get("columns", [])
    col_names = [c["name"] for c in columns]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Реестр"
    ws_out.append(["Файл"] + col_names)

    for r in rows:
        ws_out.append([r.get("_file", "")] + [r.get(n, "") for n in col_names])

    for col in ws_out.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_out.column_dimensions[col_letter].width = max_length + 2

    output_path = os.path.join(folder, "Реестр.xlsx")
    wb_out.save(output_path)
    return output_path

# ─── HTML ─────────────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Реестр документов</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=Unbounded:wght@300;700&display=swap" rel="stylesheet">
<style>
  :root {
    --bg: #0d0f14; --surface: #13161e; --surface2: #181c27;
    --border: #1e2330; --accent: #00e5a0; --accent2: #00b8ff;
    --text: #e0e6f0; --muted: #5a6480; --error: #ff4d6d;
    --warn: #ffb547;
    --mono: 'IBM Plex Mono', monospace;
    --display: 'Unbounded', sans-serif;
  }
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    background: var(--bg); color: var(--text);
    font-family: var(--mono); font-size: 13px;
    min-height: 100vh; overflow-x: hidden;
  }
  body::before {
    content: ''; position: fixed; inset: 0;
    background-image: linear-gradient(var(--border) 1px, transparent 1px), linear-gradient(90deg, var(--border) 1px, transparent 1px);
    background-size: 40px 40px; opacity: .3; pointer-events: none; z-index: 0;
  }
  .app { position: relative; z-index: 1; max-width: 1340px; margin: 0 auto; padding: 36px 24px; }

  /* ── header ── */
  header { display: flex; align-items: flex-start; justify-content: space-between; margin-bottom: 28px; flex-wrap: wrap; gap: 16px; }
  .logo { font-family: var(--display); font-weight: 700; font-size: 20px; letter-spacing: -.5px;
    background: linear-gradient(90deg, var(--accent), var(--accent2)); -webkit-background-clip: text;
    -webkit-text-fill-color: transparent; background-clip: text; }
  .logo span { font-weight: 300; }
  .subtitle { color: var(--muted); margin-top: 4px; font-size: 11px; letter-spacing: .08em; text-transform: uppercase; }

  /* ── tabs ── */
  .tabs { display: flex; gap: 2px; margin-bottom: 20px; }
  .tab {
    font-family: var(--display); font-size: 10px; font-weight: 700;
    letter-spacing: .1em; text-transform: uppercase;
    padding: 10px 20px; border: 1px solid var(--border);
    background: transparent; color: var(--muted); cursor: pointer;
    border-radius: 3px 3px 0 0; border-bottom: none;
    transition: color .15s, background .15s;
  }
  .tab:hover { color: var(--text); }
  .tab.active { color: var(--accent); background: var(--surface); border-color: var(--border); }
  .tab-line { border-bottom: 1px solid var(--border); margin-bottom: 20px; }

  .tab-content { display: none; }
  .tab-content.active { display: block; }

  /* ── panel ── */
  .panel { background: var(--surface); border: 1px solid var(--border); border-radius: 4px; padding: 20px 24px; margin-bottom: 16px; }
  .panel-label { font-family: var(--display); font-size: 10px; font-weight: 700;
    letter-spacing: .12em; text-transform: uppercase; color: var(--accent); margin-bottom: 14px; }

  /* ── inputs ── */
  .path-row { display: flex; gap: 10px; align-items: stretch; }
  input[type=text], select {
    background: var(--bg); border: 1px solid var(--border); border-radius: 3px;
    color: var(--text); font-family: var(--mono); font-size: 13px;
    padding: 9px 12px; outline: none; transition: border-color .2s;
  }
  input[type=text]:focus, select:focus { border-color: var(--accent2); }
  input[type=text]::placeholder { color: var(--muted); }
  select { cursor: pointer; }
  select option { background: var(--surface); }

  .btn {
    background: transparent; border: 1px solid var(--accent); color: var(--accent);
    font-family: var(--mono); font-size: 12px; font-weight: 600;
    letter-spacing: .06em; text-transform: uppercase;
    padding: 9px 18px; border-radius: 3px; cursor: pointer;
    transition: background .18s, color .18s; white-space: nowrap;
  }
  .btn:hover { background: var(--accent); color: var(--bg); }
  .btn:disabled { opacity: .35; cursor: default; pointer-events: none; }
  .btn.blue  { border-color: var(--accent2); color: var(--accent2); }
  .btn.blue:hover  { background: var(--accent2); color: var(--bg); }
  .btn.red   { border-color: var(--error); color: var(--error); }
  .btn.red:hover   { background: var(--error); color: var(--bg); }
  .btn.small { padding: 5px 10px; font-size: 11px; }

  /* ── stats ── */
  .stats-row { display: flex; gap: 28px; flex-wrap: wrap; }
  .stat { display: flex; flex-direction: column; gap: 2px; }
  .stat-value { font-family: var(--display); font-size: 22px; font-weight: 700; color: var(--accent); }
  .stat-label { font-size: 10px; text-transform: uppercase; letter-spacing: .08em; color: var(--muted); }

  /* ── search ── */
  .search-wrap { position: relative; margin-bottom: 14px; }
  .search-icon { position: absolute; left: 12px; top: 50%; transform: translateY(-50%); color: var(--muted); font-size: 14px; pointer-events: none; }
  .search-input { width: 100%; padding: 9px 14px 9px 34px; }

  /* ── table ── */
  .table-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; }
  thead th {
    text-align: left; font-family: var(--display); font-size: 9px; font-weight: 700;
    letter-spacing: .14em; text-transform: uppercase; color: var(--muted);
    padding: 8px 10px; border-bottom: 1px solid var(--border);
    cursor: pointer; user-select: none; white-space: nowrap;
  }
  thead th:hover { color: var(--text); }
  thead th.active { color: var(--accent); }
  tbody tr { border-bottom: 1px solid var(--border); transition: background .1s; }
  tbody tr:hover { background: rgba(0,229,160,.04); }
  tbody tr.has-error { background: rgba(255,77,109,.04); }
  tbody tr:last-child { border-bottom: none; }
  td { padding: 9px 10px; vertical-align: middle; }
  td.file { color: var(--accent2); font-size: 12px; max-width: 220px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  td.number { text-align: right; font-weight: 600; color: var(--accent); }
  td.date { color: var(--muted); }
  td.empty { color: var(--error); font-size: 11px; }
  .mark { background: rgba(0,229,160,.25); border-radius: 2px; padding: 0 1px; }

  /* ── config table ── */
  .cfg-table { width: 100%; border-collapse: collapse; margin-bottom: 14px; }
  .cfg-table th {
    text-align: left; font-family: var(--display); font-size: 9px; font-weight: 700;
    letter-spacing: .12em; text-transform: uppercase; color: var(--muted);
    padding: 6px 8px; border-bottom: 1px solid var(--border);
  }
  .cfg-table td { padding: 6px 6px; border-bottom: 1px solid rgba(30,35,48,.7); vertical-align: middle; }
  .cfg-table tr:last-child td { border-bottom: none; }
  .cfg-table input[type=text] { width: 100%; padding: 6px 8px; font-size: 12px; }
  .cfg-table select { padding: 6px 8px; font-size: 12px; }
  .cfg-table input[type=checkbox] { accent-color: var(--accent); width: 15px; height: 15px; cursor: pointer; }
  .drag-handle { color: var(--muted); cursor: grab; padding: 0 6px; font-size: 16px; user-select: none; }
  .drag-handle:active { cursor: grabbing; }
  .cfg-row.dragging { opacity: .4; }
  .cfg-row.drag-over { border-top: 2px solid var(--accent); }

  /* ── errors ── */
  .error-item { display: flex; gap: 10px; align-items: flex-start; padding: 7px 0; border-bottom: 1px solid var(--border); font-size: 12px; }
  .error-item:last-child { border-bottom: none; }
  .error-file { color: var(--error); min-width: 160px; }
  .error-msg { color: var(--muted); }

  /* ── badge ── */
  .badge { display: inline-block; background: rgba(0,229,160,.12); color: var(--accent);
    font-size: 10px; padding: 2px 7px; border-radius: 10px; margin-left: 8px; vertical-align: middle; }
  .badge.red { background: rgba(255,77,109,.12); color: var(--error); }

  /* ── toast ── */
  #toast { position: fixed; bottom: 24px; right: 24px; background: var(--surface);
    border: 1px solid var(--accent); color: var(--accent); font-family: var(--mono);
    font-size: 12px; padding: 11px 18px; border-radius: 3px;
    transform: translateY(60px); opacity: 0; transition: transform .25s, opacity .25s; z-index: 100; }
  #toast.show { transform: translateY(0); opacity: 1; }
  #toast.err  { border-color: var(--error); color: var(--error); }

  .spinner { display: inline-block; width: 12px; height: 12px; border: 2px solid var(--border);
    border-top-color: var(--accent); border-radius: 50%; animation: spin .6s linear infinite; vertical-align: middle; margin-right: 5px; }
  @keyframes spin { to { transform: rotate(360deg); } }
  .hidden { display: none !important; }

  .add-row { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; margin-top: 4px; }
  .add-row input[type=text] { flex: 1; min-width: 120px; }
  .hint { font-size: 11px; color: var(--muted); margin-top: 8px; line-height: 1.6; }

  .cfg-actions { display: flex; gap: 10px; flex-wrap: wrap; }
</style>
</head>
<body>
<div class="app">

  <header>
    <div>
      <div class="logo">Реестр<span>.</span>xlsx</div>
      <div class="subtitle">Универсальный сборщик реестров из Excel-документов</div>
    </div>
  </header>

  <div class="tabs">
    <button class="tab active" onclick="switchTab('registry')">▦ Реестр</button>
    <button class="tab" onclick="switchTab('settings')">⚙ Настройки</button>
  </div>
  <div class="tab-line"></div>

  <!-- ══ РЕЕСТР ══ -->
  <div id="tab-registry" class="tab-content active">

    <div class="panel">
      <div class="panel-label">Папка с документами</div>
      <div class="path-row">
        <input id="folderPath" type="text" style="flex:1" placeholder="Например: C:\Документы\Счета или /home/user/invoices" />
        <button class="btn" id="scanBtn" onclick="scanFolder()">▶ Сканировать</button>
        <button class="btn blue" id="saveBtn" onclick="saveXlsx()" disabled>↓ Сохранить xlsx</button>
      </div>
    </div>

    <div class="panel hidden" id="statsPanel">
      <div class="panel-label">Итого</div>
      <div class="stats-row">
        <div class="stat"><div class="stat-value" id="statFiles">0</div><div class="stat-label">Файлов</div></div>
        <div class="stat"><div class="stat-value" id="statErrors">0</div><div class="stat-label">С ошибками</div></div>
        <div class="stat"><div class="stat-value" id="statEmpty">0</div><div class="stat-label">Пустых полей</div></div>
      </div>
    </div>

    <div class="panel hidden" id="tablePanel">
      <div class="panel-label">
        Реестр <span class="badge" id="rowCount">0</span>
      </div>
      <div class="search-wrap">
        <span class="search-icon">⌕</span>
        <input class="search-input" id="searchInput" type="text" placeholder="Быстрый поиск по таблице…" oninput="filterTable()">
      </div>
      <div class="table-wrap">
        <table>
          <thead id="tableHead"><tr></tr></thead>
          <tbody id="tableBody"></tbody>
        </table>
      </div>
    </div>

    <div class="panel hidden" id="errorsPanel">
      <div class="panel-label" style="color:var(--error)">Ошибки / предупреждения <span class="badge red" id="errCount">0</span></div>
      <div id="errorList"></div>
    </div>

  </div>

  <!-- ══ НАСТРОЙКИ ══ -->
  <div id="tab-settings" class="tab-content">

    <div class="panel">
      <div class="panel-label">Колонки реестра</div>
      <p class="hint" style="margin-bottom:14px">
        Укажите для каждой колонки: название в реестре, адрес ячейки в Excel (например <b>E2</b>, <b>L12</b>),
        тип значения и нужна ли проверка на заполненность.
        Порядок колонок можно менять перетаскиванием.
      </p>
      <table class="cfg-table">
        <thead>
          <tr>
            <th style="width:28px"></th>
            <th>Название колонки</th>
            <th style="width:80px">Ячейка</th>
            <th style="width:110px">Тип</th>
            <th style="width:90px;text-align:center">Обязательное</th>
            <th style="width:50px"></th>
          </tr>
        </thead>
        <tbody id="cfgBody"></tbody>
      </table>

      <div class="add-row">
        <input type="text" id="newName" placeholder="Название" style="max-width:200px" />
        <input type="text" id="newCell" placeholder="Ячейка (E2)" style="max-width:100px;text-transform:uppercase" />
        <select id="newType">
          <option value="text">Текст</option>
          <option value="date">Дата</option>
          <option value="number">Число</option>
        </select>
        <button class="btn small" onclick="addColumn()">+ Добавить</button>
      </div>
    </div>

    <div class="panel">
      <div class="panel-label">Сохранение</div>
      <div class="cfg-actions">
        <button class="btn" onclick="saveSettings()">💾 Сохранить настройки</button>
        <button class="btn red" onclick="resetSettings()">↺ Сбросить к стандартным</button>
      </div>
      <p class="hint" style="margin-top:10px">Настройки сохраняются в <b>config.json</b> рядом со скриптом и применяются при следующем сканировании.</p>
    </div>

  </div>

</div>
<div id="toast"></div>

<script>
let allRows = [], colDefs = [], sortCol = -1, sortAsc = true;
let dragSrc = null;

// ── Tabs ──────────────────────────────────────────────────────────────────────
function switchTab(name) {
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(el => el.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  event.target.classList.add('active');
  if (name === 'settings') loadSettingsUI();
}

// ── Settings UI ───────────────────────────────────────────────────────────────
async function loadSettingsUI() {
  const res = await fetch('/api/config');
  const cfg = await res.json();
  colDefs = cfg.columns || [];
  renderCfgTable();
}

function renderCfgTable() {
  const body = document.getElementById('cfgBody');
  body.innerHTML = colDefs.map((c, i) => `
    <tr class="cfg-row" draggable="true" data-idx="${i}"
      ondragstart="dragStart(event,${i})" ondragover="dragOver(event,${i})"
      ondrop="dragDrop(event,${i})" ondragleave="dragLeave(event)">
      <td><span class="drag-handle">⠿</span></td>
      <td><input type="text" value="${esc(c.name)}" oninput="colDefs[${i}].name=this.value" /></td>
      <td><input type="text" value="${esc(c.cell)}" style="text-transform:uppercase"
            oninput="colDefs[${i}].cell=this.value.toUpperCase()" /></td>
      <td>
        <select onchange="colDefs[${i}].type=this.value">
          ${['text','date','number'].map(t => `<option value="${t}" ${c.type===t?'selected':''}>${{text:'Текст',date:'Дата',number:'Число'}[t]}</option>`).join('')}
        </select>
      </td>
      <td style="text-align:center">
        <input type="checkbox" ${c.required?'checked':''} onchange="colDefs[${i}].required=this.checked" />
      </td>
      <td><button class="btn red small" onclick="removeCol(${i})">✕</button></td>
    </tr>
  `).join('');
}

function addColumn() {
  const name = document.getElementById('newName').value.trim();
  const cell = document.getElementById('newCell').value.trim().toUpperCase();
  const type = document.getElementById('newType').value;
  if (!name || !cell) { showToast('Укажите название и ячейку', true); return; }
  colDefs.push({ name, cell, type, required: false });
  document.getElementById('newName').value = '';
  document.getElementById('newCell').value = '';
  renderCfgTable();
}

function removeCol(i) {
  colDefs.splice(i, 1);
  renderCfgTable();
}

async function saveSettings() {
  const res = await fetch('/api/config', {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({ columns: colDefs })
  });
  const data = await res.json();
  if (data.ok) showToast('Настройки сохранены');
  else showToast('Ошибка сохранения', true);
}

async function resetSettings() {
  const res = await fetch('/api/config/reset', { method: 'POST' });
  const data = await res.json();
  colDefs = data.columns || [];
  renderCfgTable();
  showToast('Настройки сброшены');
}

// drag-and-drop
function dragStart(e, i) { dragSrc = i; e.currentTarget.classList.add('dragging'); }
function dragOver(e, i) {
  e.preventDefault();
  document.querySelectorAll('.cfg-row').forEach(r => r.classList.remove('drag-over'));
  if (i !== dragSrc) e.currentTarget.classList.add('drag-over');
}
function dragLeave(e) { e.currentTarget.classList.remove('drag-over'); }
function dragDrop(e, i) {
  e.preventDefault();
  if (dragSrc === null || dragSrc === i) return;
  const moved = colDefs.splice(dragSrc, 1)[0];
  colDefs.splice(i, 0, moved);
  dragSrc = null;
  renderCfgTable();
}

// ── Registry ──────────────────────────────────────────────────────────────────
async function scanFolder() {
  const path = document.getElementById('folderPath').value.trim();
  if (!path) { showToast('Укажите путь к папке', true); return; }
  const btn = document.getElementById('scanBtn');
  btn.disabled = true; btn.innerHTML = '<span class="spinner"></span>Сканирую…';
  try {
    const res = await fetch('/api/scan?path=' + encodeURIComponent(path));
    const data = await res.json();
    if (data.error) { showToast(data.error, true); return; }
    allRows = data.rows;
    buildTableHeader(data.columns);
    filterTable();
    renderStats(data);
    renderErrors(data.errors || []);
    document.getElementById('saveBtn').disabled = false;
    showToast('Найдено файлов: ' + data.rows.length);
  } catch(e) { showToast('Ошибка соединения', true); }
  finally { btn.disabled = false; btn.innerHTML = '▶ Сканировать'; }
}

async function saveXlsx() {
  const path = document.getElementById('folderPath').value.trim();
  const btn = document.getElementById('saveBtn');
  btn.disabled = true; btn.innerHTML = '<span class="spinner"></span>Сохраняю…';
  try {
    const res = await fetch('/api/save', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({ path, rows: allRows })
    });
    const data = await res.json();
    if (data.error) { showToast(data.error, true); return; }
    showToast('Сохранено: ' + data.output_path);
  } catch(e) { showToast('Ошибка сохранения', true); }
  finally { btn.disabled = false; btn.innerHTML = '↓ Сохранить xlsx'; }
}

function buildTableHeader(columns) {
  const tr = document.querySelector('#tableHead tr');
  const cols = ['Файл', ...columns];
  tr.innerHTML = cols.map((c, i) =>
    `<th onclick="sortTable(${i})" data-col="${i}">${esc(c)} <span class="sort-arrow">↕</span></th>`
  ).join('');
}

function renderStats(data) {
  let emptyCount = 0;
  data.rows.forEach(r => {
    data.columns.forEach(c => { if (r[c] === null || r[c] === undefined || r[c] === '') emptyCount++; });
  });
  document.getElementById('statFiles').textContent  = data.rows.length;
  document.getElementById('statErrors').textContent = (data.errors||[]).length;
  document.getElementById('statEmpty').textContent  = emptyCount;
  document.getElementById('statsPanel').classList.remove('hidden');
}

function renderErrors(errors) {
  const panel = document.getElementById('errorsPanel');
  document.getElementById('errCount').textContent = errors.length;
  if (!errors.length) { panel.classList.add('hidden'); return; }
  panel.classList.remove('hidden');
  document.getElementById('errorList').innerHTML = errors.map(e =>
    `<div class="error-item"><span class="error-file">${esc(e.file)}</span><span class="error-msg">${esc(e.error)}</span></div>`
  ).join('');
}

function renderTable(rows) {
  const q = document.getElementById('searchInput').value.toLowerCase();
  document.getElementById('tablePanel').classList.remove('hidden');
  const cols = Array.from(document.querySelectorAll('#tableHead th')).map(th => th.dataset.col);
  const colNames = Array.from(document.querySelectorAll('#tableHead th')).map(th => th.textContent.replace(/[↕↑↓]/g,'').trim());

  const body = document.getElementById('tableBody');
  body.innerHTML = rows.map(r => {
    const file = r['_file'] || '';
    const cells = colNames.map(c => c === 'Файл' ? file : r[c]);
    return `<tr>
      ${cells.map((v, i) => {
        const raw = v === null || v === undefined ? '' : String(v);
        const isEmpty = raw === '';
        const col = colNames[i];
        let cls = '';
        if (i === 0) cls = 'file';
        else if (typeof v === 'number') cls = 'number';
        else if (col === 'Дата' || (raw && raw.match(/^\d{4}-\d{2}-\d{2}$/))) cls = 'date';
        if (isEmpty) cls += ' empty';
        const display = isEmpty ? '—' : (typeof v === 'number' ? v.toLocaleString('ru-RU', {maximumFractionDigits:2}) : raw);
        return `<td class="${cls.trim()}">${hi(esc(display), q)}</td>`;
      }).join('')}
    </tr>`;
  }).join('');

  document.getElementById('rowCount').textContent = rows.length;
}

function filterTable() {
  const q = document.getElementById('searchInput').value.toLowerCase();
  const filtered = !q ? allRows : allRows.filter(r =>
    Object.values(r).some(v => v !== null && String(v).toLowerCase().includes(q))
  );
  renderTable(filtered);
}

function sortTable(col) {
  const ths = document.querySelectorAll('#tableHead th');
  const colNames = Array.from(ths).map(th => th.textContent.replace(/[↕↑↓]/g,'').trim());
  if (sortCol === col) sortAsc = !sortAsc; else { sortCol = col; sortAsc = true; }
  ths.forEach((th, i) => {
    th.classList.toggle('active', i === col);
    th.querySelector('.sort-arrow').textContent = i === col ? (sortAsc ? '↑' : '↓') : '↕';
  });
  const key = col === 0 ? '_file' : colNames[col];
  allRows.sort((a, b) => {
    let av = a[key], bv = b[key];
    if (typeof av === 'number' && typeof bv === 'number') return sortAsc ? av - bv : bv - av;
    av = String(av||'').toLowerCase(); bv = String(bv||'').toLowerCase();
    return sortAsc ? av.localeCompare(bv) : bv.localeCompare(av);
  });
  filterTable();
}

function hi(text, q) {
  if (!q || !text) return text;
  const idx = text.toLowerCase().indexOf(q);
  if (idx < 0) return text;
  return text.slice(0,idx) + '<mark class="mark">' + text.slice(idx,idx+q.length) + '</mark>' + text.slice(idx+q.length);
}
function esc(s) { return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

let toastTimer;
function showToast(msg, isErr=false) {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className = 'show' + (isErr ? ' err' : '');
  clearTimeout(toastTimer);
  toastTimer = setTimeout(() => { t.className = ''; }, 3500);
}

document.getElementById('folderPath').addEventListener('keydown', e => { if (e.key==='Enter') scanFolder(); });
</script>
</body>
</html>
"""

# ─── HTTP Handler ─────────────────────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args): pass

    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path in ("/", "/index.html"):
            self._html(HTML)

        elif parsed.path == "/api/scan":
            qs = parse_qs(parsed.query)
            folder = unquote(qs.get("path", [""])[0]).strip()
            if not folder or not os.path.isdir(folder):
                self._json({"error": f"Папка не найдена: {folder}"}); return
            cfg = load_config()
            data = build_registry(folder, cfg)
            self._json(data)

        elif parsed.path == "/api/config":
            self._json(load_config())

        else:
            self.send_response(404); self.end_headers()

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length)

        if self.path == "/api/save":
            payload = json.loads(body)
            folder = payload.get("path","").strip()
            rows   = payload.get("rows", [])
            if not folder or not os.path.isdir(folder):
                self._json({"error": f"Папка не найдена: {folder}"}); return
            try:
                cfg = load_config()
                out = save_registry_xlsx(folder, rows, cfg)
                self._json({"output_path": out})
            except Exception as e:
                self._json({"error": str(e)})

        elif self.path == "/api/config":
            try:
                cfg = json.loads(body)
                save_config(cfg)
                self._json({"ok": True})
            except Exception as e:
                self._json({"error": str(e)})

        elif self.path == "/api/config/reset":
            save_config(DEFAULT_CONFIG)
            self._json(DEFAULT_CONFIG)

        else:
            self.send_response(404); self.end_headers()

    def _html(self, html):
        body = html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _json(self, obj):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


# ─── CLI mode ─────────────────────────────────────────────────────────────────

def run_cli_mode():
    folder = os.path.dirname(os.path.abspath(__file__))
    cfg = load_config()
    data = build_registry(folder, cfg)
    for e in data["errors"]:
        print(f"[ERROR] {e['file']}: {e['error']}")
    for r in data["rows"]:
        print(f"[OK] {r['_file']}")
    out = save_registry_xlsx(folder, data["rows"], cfg)
    print(f"\n✅ Реестр сформирован: {out}")


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    if "--cli" in sys.argv:
        run_cli_mode()
    else:
        PORT = 8765
        server = HTTPServer(("127.0.0.1", PORT), Handler)
        url = f"http://127.0.0.1:{PORT}"
        print(f"✅ Сервер запущен: {url}")
        print("   Остановить: Ctrl+C")
        print("   CLI-режим:  python registry_app.py --cli")
        threading.Timer(1.2, lambda: webbrowser.open(url)).start()
        try:
            server.serve_forever()
        except KeyboardInterrupt:
            print("\n⛔ Остановлено.")
